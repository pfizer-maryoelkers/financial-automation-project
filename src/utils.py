import yaml
import base64
import io
import pandas as pd
import re
from collections import defaultdict
from src.models import CostCenter, WBSCode, PO, MonthlyMetrics, ExceptionLog, ExceptionType


def detect_template_type(file_path: str) -> str:
    """Inspect the template workbook and return which pipeline it belongs to.

    Returns
    -------
    "project"
        The top corner has "WBS Code" and "P3 ID" as a **paired header row**
        (both values appear on the same row, one in col A and one in col B).
        The project pipeline (ProjectTemplateReader / build_project_hierarchy)
        should be used.
    "opex"
        The template has a "Cost Center" section in column A.  The OpEx
        pipeline (TemplateReader / build_hierarchy) should be used.

    Detection strategy
    ------------------
    Scan up to the first 20 rows.  A row where col-A says "WBS Code" (or
    "WBS code") AND col-B says "P3 ID" on the same row is an unambiguous
    project-template marker.  The OpEx template has those labels too, but
    never together on the same row.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    for r in range(1, 21):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value
        a_text = str(a_val).strip().lower() if a_val is not None else ""
        b_text = str(b_val).strip().lower() if b_val is not None else ""
        # Project template: "WBS Code" in col A AND "P3 ID" in col B on the same row
        if a_text == "wbs code" and b_text == "p3 id":
            wb.close()
            return "project"

    wb.close()
    return "opex"


def load_config(config_path='configs/config_base.yaml'):
    """Load and merge YAML configs.
    
    Defaults to values in 'config_base.yaml', overrides if there are any overrides needed
    """
    with open(config_path) as f:
        config = yaml.safe_load(f)
    
    return config

# Helper function to combine forecasts, actuals, and accrual data into one JSON formatted dictionary
def combine_data(forecast, transactional):
    combined = {}

    # All PO numbers that exist in either dataset
    all_pos = set(forecast.keys()) | set(transactional.keys())

    months_list = [
        "Jan","Feb","Mar","Apr","May","Jun",
        "Jul","Aug","Sep","Oct","Nov","Dec"
    ]

    for po in all_pos:
        combined[po] = {}

        for month in months_list:
            f = forecast.get(po, {}).get(month, {})
            t = transactional.get(po, {}).get(month, {})

            combined[po][month] = {
                "Forecast": f.get("Forecast", 0),
                "Actual": t.get("Actual", 0),
                "Accrual": t.get("Accrual", 0),
                "Accrual Reversal": t.get("Reversal", 0)
            }

    return combined

def convert_base64(bytes_string: str):
    # Converts string of bytes to Excel like object for pd/openpyxl to read
    decoded_bytes = base64.b64decode(bytes_string)
    excel_file_like_object = io.BytesIO(decoded_bytes)
    return excel_file_like_object


# Forecast month shift: maps a 3-letter month key one step back.
_FORECAST_MONTH_SHIFT = {
    "Jan": "Dec",   # Jan forecast → Dec (prior year treated as Dec in same year for intl)
    "Feb": "Jan",
    "Mar": "Feb",
    "Apr": "Mar",
    "May": "Apr",
    "Jun": "May",
    "Jul": "Jun",
    "Aug": "Jul",
    "Sep": "Aug",
    "Oct": "Sep",
    "Nov": "Oct",
    "Dec": "Nov",
}

def build_hierarchy(
    cost_centers: list[str],
    hierarchy_map: dict,
    transactional_data: dict,
    forecast_data: dict,
    exception_log: ExceptionLog,
    transactional_df: pd.DataFrame,
    reclass_data: dict | None = None,
    reclass_notes: dict | None = None,
    template_pos: dict | None = None,
    intl_po_set: set | None = None,
) -> dict[str, CostCenter]:
    
    # Regex pattern to extract ER numbers (ER followed by digits)
    er_pattern = re.compile(r'\bER\d+\b', re.IGNORECASE)

    def _find_er_in_row(row: dict) -> str | None:
        """Search GL Line Description, GL Transaction Description, and Description
        columns (in that priority order) for an ER number. Returns the first match
        uppercased (e.g. 'ER97054'), or None if not found."""
        for key in ('gl_line_desc', 'gl_trans_desc', 'description'):
            text = row.get(key)
            if text:
                m = er_pattern.search(text)
                if m:
                    return m.group(0).upper()
        return None

    # Step 1: Pre-group hierarchy_map rows by cost center
    rows_by_cost_center = defaultdict(list)
    for row_idx, row in hierarchy_map.items():
        if row['cost_center'] in cost_centers:
            rows_by_cost_center[row['cost_center']].append((row_idx, row))

    # Step 2: Pre-scan to identify duplicate WBS codes and real WBS per ER number
    wbs_to_cost_centers = defaultdict(set)  # wbs -> set of cost_centers
    er_real_wbs = {}  # er_number -> real WBS from any row that has both ER PO and a WBS
    for cc_id in cost_centers:
        for row_idx, row in rows_by_cost_center[cc_id]:
            wbs = row['wbs']
            po = row['po']
            if wbs:  # Only track non-empty WBS
                wbs_to_cost_centers[wbs].add(cc_id)
                # If the PO itself looks like an ER number and this row has a real WBS, capture it
                if po and er_pattern.match(po) and po not in er_real_wbs:
                    er_real_wbs[po] = wbs

    # Identify which WBS codes are duplicates (appear under multiple cost centers)
    duplicate_wbs_codes = {wbs for wbs, ccs in wbs_to_cost_centers.items() if len(ccs) > 1}

    # Step 3: Track seen POs and WBS codes for processing
    seen_pos = {}  # po -> (cost_center, wbs)
    seen_wbs = {}  # wbs -> cost_center (for tracking first occurrence)
    er_extracted_count = 0

    # Step 4: Build hierarchy
    result = {}
    for cc_id in cost_centers:
        cost_center = CostCenter(cost_center_id=cc_id)
        for row_idx, row in rows_by_cost_center[cc_id]:
            po = row['po']
            wbs = row['wbs']
            legal_entity = row.get('legal_entity')
            country = row.get('country')
            vendor_name = row.get('vendor_name')
            gl_account = row.get('gl_account')
            req_title = row.get('req_title')

            # Extract source row data from transactional DataFrame
            source_row_data = transactional_df.loc[row_idx].to_dict() if row_idx in transactional_df.index else {}
            month = source_row_data.get('Accounting Period')
            amount = source_row_data.get('GL BER Corp Amount')
            trans_type = source_row_data.get('Type')

            # Handling Exceptions (in priority order)

            # Check 0a: Unmatched transaction type — not Actual, Accrual, Reversal,
            # Reclass, or ER. Only log when the PO appears on the front template tab
            # so exceptions stay relevant to what's already being tracked.
            _KNOWN_TYPES = {"Actual", "Accrual", "Reversal", "Reclass", "ER"}
            if trans_type not in _KNOWN_TYPES:
                if template_pos and po and po in template_pos:
                    exception_log.log(
                        ExceptionType.UNMATCHED_TRANSACTION,
                        row_index=row_idx,
                        po=po,
                        wbs=wbs,
                        cost_center=cc_id,
                        month=month,
                        amount=amount,
                        transaction_type=trans_type,
                        source_row_data=source_row_data
                    )
                continue

            # Check 0b: Reclass rows — log as exceptions then skip. Their amount is
            # already folded into the PO's Actual total inside get_transactional_data().
            if trans_type == "Reclass":
                exception_log.log(
                    ExceptionType.RECLASS,
                    row_index=row_idx,
                    po=po,
                    wbs=wbs,
                    cost_center=cc_id,
                    month=month,
                    amount=amount,
                    transaction_type=trans_type,
                    source_row_data=source_row_data
                )
                continue

            # Check 1: Both WBS and PO missing (highest priority)
            # Special case: Try to extract ER number from any description column.
            # ER rows are already typed "ER" by _categorize_row so they never
            # reach this block with trans_type == "Reclass".
            if not wbs and not po:
                er_found = _find_er_in_row(row)
                if er_found:
                    # Use extracted ER number as PO with special WBS "ER"
                    po = er_found  # e.g., ER97054
                    wbs = "ER"    # Special WBS code for expense reports
                    er_extracted_count += 1
                    # Continue processing with extracted ER as PO
                else:
                    # No ER found — only flag when the template already has PO rows
                    if template_pos:
                        exception_log.log(
                            ExceptionType.MISSING_WBS_AND_PO,
                            row_index=row_idx,
                            cost_center=cc_id,
                            month=month,
                            amount=amount,
                            transaction_type=trans_type,
                            source_row_data=source_row_data
                        )
                    continue

            # Check 1b: PO exists but looks like a full ER description — extract clean ER number
            # e.g. PO column contains "ER97054 - ARCH & TECH - APP ARCHITECTURE"
            if po and not wbs:
                er_found = _find_er_in_row(row)
                if not er_found:
                    # Fall back: try extracting directly from the PO value itself
                    m = er_pattern.match(po)
                    if m:
                        er_found = m.group(0).upper()
                if er_found:
                    po = er_found
                    wbs = "ER"
                    er_extracted_count += 1

            # Check 2: Individual missing checks
            if not wbs:
                # Only flag when the template already has PO rows
                if template_pos:
                    exception_log.log(
                        ExceptionType.MISSING_WBS,
                        row_index=row_idx,
                        po=po,
                        cost_center=cc_id,
                        month=month,
                        amount=amount,
                        transaction_type=trans_type,
                        source_row_data=source_row_data
                    )
                # Still place the PO on the first tab under a fallback WBS bucket
                # so the data is visible even though no WBS code was found.
                wbs = "NO_WBS"
            
            if not po:
                # Only flag MISSING_PO when the template front tab already has PO rows.
                # If the template is blank there is nothing to relate the transaction to,
                # so suppressing the exception avoids noise for new/empty templates.
                if template_pos:
                    exception_log.log(
                        ExceptionType.MISSING_PO,
                        row_index=row_idx,
                        wbs=wbs,
                        cost_center=cc_id,
                        month=month,
                        amount=amount,
                        transaction_type=trans_type,
                        source_row_data=source_row_data
                    )
                continue

            # Check 3: Duplicate WBS (WBS owned by multiple cost centers)
            # Log ALL occurrences of duplicate WBS codes as exceptions
            if wbs in duplicate_wbs_codes:
                exception_log.log(
                    ExceptionType.DUPLICATE_WBS,
                    row_index=row_idx,
                    wbs=wbs,
                    po=po,
                    cost_center=cc_id,
                    month=month,
                    amount=amount,
                    transaction_type=trans_type,
                    source_row_data=source_row_data
                )
                # Still process first occurrence in hierarchy, skip subsequent ones
                if wbs in seen_wbs and seen_wbs[wbs] != cc_id:
                    continue  # Skip this occurrence - first is canonical
            
            seen_wbs[wbs] = cc_id

            # Check 4: Duplicate PO
            if po in seen_pos:
                prev_cc, prev_wbs = seen_pos[po]
                if prev_cc != cc_id or prev_wbs != wbs:
                    exception_log.log(
                        ExceptionType.DUPLICATE_PO,
                        row_index=row_idx,
                        po=po,
                        wbs=wbs,
                        cost_center=cc_id,
                        month=month,
                        amount=amount,
                        transaction_type=trans_type,
                        source_row_data=source_row_data
                    )
                continue  # Skip regardless - first occurrence is canonical

            seen_pos[po] = (cc_id, wbs)

            # Check 5: PO is in the transactional file for a tracked cost center but
            # is NOT on the front template tab. Only fires when the template already
            # has PO rows (populated template) and only logged once per PO.
            # Log every transaction row for a PO that is not on the template so
            # the exceptions tab can show per-month amounts and a full total.
            if template_pos and po not in template_pos and wbs != "ER":
                exception_log.log(
                    ExceptionType.PO_NOT_ON_TEMPLATE,
                    row_index=row_idx,
                    po=po,
                    wbs=wbs,
                    cost_center=cc_id,
                    month=month,
                    amount=amount,
                    transaction_type=trans_type,
                    vendor_name=vendor_name,
                    source_row_data=source_row_data
                )

            # Build WBS and PO objects if not already seen
            if wbs not in cost_center.wbs_codes:
                cost_center.wbs_codes[wbs] = WBSCode(wbs_code=wbs, cost_center=cc_id)
            if po not in cost_center.wbs_codes[wbs].pos:
                cost_center.wbs_codes[wbs].pos[po] = PO(
                    po_number=po,
                    legal_entity=legal_entity,
                    country=country,
                    vendor_name=vendor_name,
                    gl_account=gl_account,
                    req_title=req_title,
                    real_wbs=er_real_wbs.get(po) if wbs == "ER" else None,
                )
            po_obj = cost_center.wbs_codes[wbs].pos[po]
            # Backfill any fields that were None on the first row but have a real value now
            if po_obj.vendor_name is None and vendor_name is not None:
                po_obj.vendor_name = vendor_name
            if po_obj.legal_entity is None and legal_entity is not None:
                po_obj.legal_entity = legal_entity
            if po_obj.country is None and country is not None:
                po_obj.country = country
            if po_obj.gl_account is None and gl_account is not None:
                po_obj.gl_account = gl_account
            if po_obj.req_title is None and req_title is not None:
                po_obj.req_title = req_title
            
            # Fill MonthlyMetrics from transactional data
            po_lookup = str(po).strip().upper() if wbs == "ER" and po else po
            if po_lookup in transactional_data:
                po_data = transactional_data[po_lookup]
                # Capture gross BER total for Gross PO Value column
                if po_obj.gross_po_value is None:
                    po_obj.gross_po_value = po_data.get('gross_ber_total')
                for month, values in po_data.items():
                    if month in ('cost_center', 'wbs', 'gross_ber_total'):
                        continue
                    if month not in po_obj.monthly_data:
                        po_obj.monthly_data[month] = MonthlyMetrics()
                    metrics = po_obj.monthly_data[month]
                    metrics.actual = values.get('Actual', 0.0)
                    metrics.accrual = values.get('Accrual', 0.0)
                    metrics.accrual_reversal = values.get('Reversal', 0.0)
            
            # Fill forecast from forecast data (only for non-zero values).
            # International POs shift the forecast month back by 1 (same rule as
            # their transactional data) so Forecast and Actual land in the same column.
            if po in forecast_data:
                is_intl_po = bool(intl_po_set and str(po).strip() in intl_po_set)
                for month, values in forecast_data[po].items():
                    fc_value = values.get('Forecast', 0.0)
                    if not fc_value:  # skip 0 / None — leave monthly_data entry absent
                        continue
                    write_month = _FORECAST_MONTH_SHIFT.get(month, month) if is_intl_po else month
                    if write_month not in po_obj.monthly_data:
                        po_obj.monthly_data[write_month] = MonthlyMetrics()
                    po_obj.monthly_data[write_month].forecast = fc_value
            # Note: MISSING_FORECAST exception removed - no longer tracking

        # Attach reclass notes once per PO after the row loop, so entries are
        # never duplicated regardless of how many transactional rows a PO has.
        # Also ensures POs that appear ONLY in Reclass rows still get built into
        # the hierarchy (their Actual total is already in transactional_data).
        if reclass_notes:
            for po_number, month_entries in reclass_notes.items():
                # Only process POs that belong to this cost center
                po_in_td = transactional_data.get(po_number, {})
                if po_in_td.get('cost_center') != cc_id:
                    continue
                # Find or create the WBS + PO object
                wbs_key = po_in_td.get('wbs', 'NO_WBS') or 'NO_WBS'
                if wbs_key not in cost_center.wbs_codes:
                    cost_center.wbs_codes[wbs_key] = WBSCode(wbs_code=wbs_key, cost_center=cc_id)
                if po_number not in cost_center.wbs_codes[wbs_key].pos:
                    # PO existed only as Reclass rows — build it now from transactional_data
                    cost_center.wbs_codes[wbs_key].pos[po_number] = PO(po_number=po_number)
                    po_obj = cost_center.wbs_codes[wbs_key].pos[po_number]
                    for month, values in po_in_td.items():
                        if month in ('cost_center', 'wbs', 'gross_ber_total'):
                            continue
                        if month not in po_obj.monthly_data:
                            po_obj.monthly_data[month] = MonthlyMetrics()
                        po_obj.monthly_data[month].actual = values.get('Actual', 0.0)
                        po_obj.monthly_data[month].accrual = values.get('Accrual', 0.0)
                        po_obj.monthly_data[month].accrual_reversal = values.get('Reversal', 0.0)
                else:
                    po_obj = cost_center.wbs_codes[wbs_key].pos[po_number]
                # Attach notes — use a set to avoid duplicates
                for month_label, entries in month_entries.items():
                    if month_label not in po_obj.reclass_adjustments:
                        po_obj.reclass_adjustments[month_label] = list(entries)
                    else:
                        existing = set(po_obj.reclass_adjustments[month_label])
                        for entry in entries:
                            if entry not in existing:
                                po_obj.reclass_adjustments[month_label].append(entry)
                                existing.add(entry)

        result[cc_id] = cost_center

    if er_extracted_count > 0:
        print(f"  - ER numbers extracted and processed: {er_extracted_count}")

    return result