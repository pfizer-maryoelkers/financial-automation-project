from collections import defaultdict
from copy import copy
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models import CostCenter, WBSCode, PO, MonthlyMetrics, ExceptionLog, ExceptionType


def _copy_color(c: Color) -> Color:
    """Deep-copy a Color that may be rgb, theme, or indexed.
    openpyxl's copy() drops theme/tint on theme-based colours."""
    if c is None:
        return Color()
    if c.type == 'theme':
        new = Color(theme=c.theme, tint=c.tint)
    elif c.type == 'indexed':
        new = Color(indexed=c.indexed)
    else:
        # rgb or auto
        new = Color(rgb=c.rgb)
    return new


def _copy_fill(f: PatternFill) -> PatternFill:
    """Deep-copy a PatternFill, preserving theme-based colours."""
    if f is None or f.fill_type is None:
        return PatternFill()
    return PatternFill(
        fill_type=f.fill_type,
        fgColor=_copy_color(f.fgColor),
        bgColor=_copy_color(f.bgColor),
    )



def _append_comment(existing_comment: Comment | None, new_text: str) -> Comment:
    """Preserve any existing note/comment text and append the new text."""
    if existing_comment and existing_comment.text:
        combined_text = f"{existing_comment.text}\n\n{new_text}"
    else:
        combined_text = new_text

    comment = Comment(combined_text, "Financial Automation")
    line_count = len(combined_text.split('\n'))
    comment.width = 400
    comment.height = max(120, 60 + 20 * line_count)
    return comment


def month_sort_key(month_str):
    """Convert month string to sortable key for proper chronological ordering"""
    month_order = {
        'Dec (PY)': 0,
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12,
        'Unknown': 13
    }
    return month_order.get(month_str, 13)



class TemplateWriter:

    def __init__(
        self,
        file_path,
        output_path,
        overwrite,
        header_row,
        po_column,
        dec_acc_reversal_col,
        forecast_source_cols,
        transactional_source_cols,
        p3_id_column: str | None = None,
    ):
        self.wb = load_workbook(file_path)
        self.sheet: Worksheet = self.wb.active  # type: ignore[assignment]
        if self.sheet is None:
            raise ValueError(
                f"Could not load active sheet from the template file '{file_path}'. "
                f"Please verify that the file exists, is not corrupted, and contains a valid active worksheet. "
                f"Available sheets: {self.wb.sheetnames}"
            )

        self.output_path = output_path
        self.overwrite = overwrite
        self.po_column = po_column

        # Dynamically locate the header row by scanning the sheet; the config
        # value is only used as a fallback when no header marker is found.
        self.header_row = header_row
        self.header_row = self._find_actual_header_row()

        # Optional P3-ID column (project pipeline). None → OpEx default (col J).
        self.p3_id_column: int | None = (
            column_index_from_string(p3_id_column) if p3_id_column else None
        )

        # Month → metric → column-letter map built from the header row.
        self.dec_acc_reversal_col = dec_acc_reversal_col
        self.column_map = self.get_column_map(starting_col=self.dec_acc_reversal_col)

        self.forecast_source_cols = forecast_source_cols
        self.transactional_source_cols = transactional_source_cols
        self.forecast_po_col = self.forecast_source_cols[0]
        self.transactional_po_col = self.transactional_source_cols[0]

        # Columns discovered dynamically from the header row.
        self.total_col = self._get_total_col()
        self.po_value_col = self._get_po_value_col()
        self.req_title_cols = self._get_req_title_cols()
        self.project_name_col = self._get_col_by_header("project name")
        self.po_status_col = self._get_col_by_header("po status")

    @staticmethod
    def _norm_po(v) -> str:
        """Normalise a PO value to a clean string (strip .0 from float-integers)."""
        s = str(v).strip()
        if s.replace('.', '', 1).replace('-', '', 1).isdigit():
            try:
                return str(int(float(s)))
            except (ValueError, OverflowError):
                pass
        return s

    def _should_write(self, existing) -> bool:
        """Return True when a cell value should be overwritten.

        Rules:
        - Always write when overwrite=True.
        - Never overwrite an existing formula when overwrite=False.
        - Write when the cell is blank (None or empty string).
        """
        if self.overwrite:
            return True
        if isinstance(existing, str) and existing.startswith('='):
            return False
        return existing is None or str(existing).strip() == ""

    def _get_po_value_col(self) -> int | None:
        """Scan the header row for the column that receives invoice amount values.
        Returns the 'invoice amount' column only; PO Total / Gross PO stays blank."""
        for col in range(1, (self.sheet.max_column or 200) + 1):
            val = self.sheet.cell(row=self.header_row, column=col).value
            if not val:
                continue
            text = str(val).strip().lower()
            if "invoice amount" in text:
                return col
        return None

    def _get_req_title_cols(self) -> list[int]:
        """Scan the header row for all requisition/description columns.
        Returns a list of column indices (1-based); may be empty."""
        found = []
        for col in range(1, (self.sheet.max_column or 200) + 1):
            val = self.sheet.cell(row=self.header_row, column=col).value
            if not val:
                continue
            text = str(val).strip().lower()
            if "requisition" in text or "gl transaction description" in text:
                found.append(col)
        return found

    def _get_col_by_header(self, keyword: str) -> int | None:
        """Scan the header row for a column whose label contains keyword (case-insensitive)."""
        for col in range(1, (self.sheet.max_column or 200) + 1):
            val = self.sheet.cell(row=self.header_row, column=col).value
            if val and keyword.lower() in str(val).lower():
                return col
        return None

    def _get_total_col(self) -> str | None:
        """Scan a window around the header row for a 'Total YYYY' cell.
        Returns the column letter, or None if not found."""
        for row in range(1, self.header_row + 10):
            for col in range(1, (self.sheet.max_column or 200) + 1):
                val = self.sheet.cell(row=row, column=col).value
                if val and re.search(r'total\s+\d{4}', str(val), re.IGNORECASE):
                    return get_column_letter(col)
        return None

    def _write_total_formula(self, row: int):
        """Write the Total 2026 SUM formula into the total column for the given row.
        The formula picks the best available value per month:
        Actual → Accrual → Forecast (matching the pattern in the template).
        Only writes if the total column was found and the cell is blank (or overwrite=True).
        Preserves any formula already in the cell."""
        if not self.total_col:
            return
        cell = self.sheet[f"{self.total_col}{row}"]
        existing = cell.value
        # Preserve existing formula in this cell
        if not self.overwrite and isinstance(existing, str) and existing.startswith('='):
            return
        if not self.overwrite and existing not in (None, 0, ''):
            return
        # Build SUM of per-month best-value: IF(Actual<>"", Actual, IF(Accrual<>"", Accrual, Forecast))
        parts = []
        for month_cols in self.column_map.values():
            actual   = month_cols.get('Actual')
            accrual  = month_cols.get('Accrual')
            forecast = month_cols.get('Forecast')
            if actual and accrual and forecast:
                parts.append(
                    f'IF({actual}{row}<>"",{actual}{row},'
                    f'IF({accrual}{row}<>"",{accrual}{row},{forecast}{row}))'
                )
        if parts:
            cell.value = '=' + '+'.join(parts)

    def _update_summary_formulas(self):
        """Rewrite range references in summary rows (rows 1 … header_row-1) so
        they span from header_row+1 to the actual last data row.

        The template ships with hardcoded ranges like K17:K18 (only 2 placeholder
        rows).  After PO rows are inserted those ranges need updating to cover all
        real data rows, otherwise SUBTOTAL / SUM / SUMIFS totals show wrong values.

        Strategy: for every formula cell above the header row, replace every
        occurrence of <COL><start_row>:<COL><end_row> where start_row equals
        header_row+1 with the correct end row.  Handles plain string formulas
        and ArrayFormula objects.
        """
        from openpyxl.worksheet.formula import ArrayFormula
        import re

        data_first = self.header_row + 1  # first real data row (e.g. 17)

        # Find the actual last data row (row before 'EXPENSE END' which includes 'Previous Period Invoices')
        data_last = data_first
        for r in range(data_first, (self.sheet.max_row or 1000) + 1):
            val = self.sheet.cell(row=r, column=1).value
            if val is not None and str(val).strip() == "EXPENSE END":
                data_last = r - 1
                break
            data_last = r

        if data_last < data_first:
            return  # nothing to do (empty template)

        def _replace_range(formula_text: str) -> str:
            """Replace <COL><data_first>:<COL><old_end> with <COL><data_first>:<COL><data_last>."""
            def replacer(m):
                col1, row1, col2, row2 = m.group(1), m.group(2), m.group(3), m.group(4)
                if int(row1) == data_first:
                    full_match = m.group(0)
                    has_dollar_col1 = full_match.startswith('$')
                    part1, part2 = full_match.split(':')
                    has_dollar_row1 = '$' in part1[1:] if has_dollar_col1 else '$' in part1
                    has_dollar_col2 = part2.startswith('$')
                    has_dollar_row2 = '$' in part2[1:] if has_dollar_col2 else '$' in part2

                    d_col1 = '$' if has_dollar_col1 else ''
                    d_row1 = '$' if has_dollar_row1 else ''
                    d_col2 = '$' if has_dollar_col2 else ''
                    d_row2 = '$' if has_dollar_row2 else ''

                    return f"{d_col1}{col1}{d_row1}{row1}:{d_col2}{col2}{d_row2}{data_last}"
                return m.group(0)
            return re.sub(r'\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)', replacer, formula_text)

        # Build dynamic mapping from summary column index to its correct accrual column letter
        summary_cols = {}
        if self.p3_id_column is not None:
            if 'Dec (PY)' in self.column_map and 'Accrual Reversal' in self.column_map['Dec (PY)']:
                col_letter = self.column_map['Dec (PY)']['Accrual Reversal']
                jan_acc = self.column_map.get('Jan', {}).get('Accrual')
                if jan_acc:
                    summary_cols[column_index_from_string(col_letter)] = jan_acc

            months_ordered = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            for idx, m_key in enumerate(months_ordered[:-1]):
                next_month = months_ordered[idx + 1]
                if m_key in self.column_map and 'Accrual Reversal' in self.column_map[m_key]:
                    col_letter = self.column_map[m_key]['Accrual Reversal']
                    next_acc = self.column_map.get(next_month, {}).get('Accrual')
                    if next_acc:
                        summary_cols[column_index_from_string(col_letter)] = next_acc

        def _correct_accrual_column(formula_text: str, col_idx: int) -> str:
            if col_idx not in summary_cols:
                return formula_text

            correct_col = summary_cols[col_idx]

            # Replace references to column L ($L or L) with correct_col
            # specifically within ranges matching \$?L\$?(\d+):\$?L\$?(\d+)
            def repl(m):
                full_match = m.group(0)
                has_dollar_col1 = full_match.startswith('$')
                parts = full_match.split(':')
                has_dollar_col2 = parts[1].startswith('$')
                has_dollar_row1 = '$' in parts[0][1:] if has_dollar_col1 else '$' in parts[0]
                has_dollar_row2 = '$' in parts[1][1:] if has_dollar_col2 else '$' in parts[1]

                d_col1 = '$' if has_dollar_col1 else ''
                d_row1 = '$' if has_dollar_row1 else ''
                d_col2 = '$' if has_dollar_col2 else ''
                d_row2 = '$' if has_dollar_row2 else ''

                row1 = m.group(1)
                row2 = m.group(2)

                return f"{d_col1}{correct_col}{d_row1}{row1}:{d_col2}{correct_col}{d_row2}{row2}"

            return re.sub(r'\$?L\$?(\d+):\$?L\$?(\d+)', repl, formula_text)

        updated = 0
        for row in range(1, self.header_row):
            for col in range(1, (self.sheet.max_column or 200) + 1):
                cell = self.sheet.cell(row=row, column=col)
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, ArrayFormula):
                    new_text = _replace_range(v.text)
                    new_text = _correct_accrual_column(new_text, col)
                    if new_text != v.text:
                        cell.value = ArrayFormula(v.ref, new_text)
                        updated += 1
                elif isinstance(v, str) and v.startswith('='):
                    new_formula = _replace_range(v)
                    new_formula = _correct_accrual_column(new_formula, col)
                    if new_formula != v:
                        cell.value = new_formula
                        updated += 1

        if updated:
            print(f"Updated {updated} summary formula(s) to cover rows "
                  f"{data_first}:{data_last}.")

        # Dynamically align 'Previous Period Invoices' row formulas
        self._update_previous_period_invoices_formulas()

    def _update_previous_period_invoices_formulas(self):
        """Update any formulas on the 'Previous Period Invoices' row so they
        refer to the correct row index (after row insertions). Also repairs
        any missing monthly Variance and Total formulas on this row.
        """
        ppi_row = None
        for r in range(self.header_row + 1, (self.sheet.max_row or 1000) + 1):
            val = self.sheet.cell(row=r, column=1).value
            if val is not None and str(val).strip() == "Previous Period Invoices":
                ppi_row = r
                break
                
        if ppi_row is None:
            return  # Previous Period Invoices row not found
            
        # 1. Update/Rewrite the monthly Variance formulas
        # We can loop through each month in column_map, get the Accrual and Actual columns,
        # and write the IF(OR(Accrual="",Actual=""),"",Accrual-Actual) formula into the Variance column.
        for month_key, month_cols in self.column_map.items():
            if month_key == "Dec (PY)":
                continue
            actual_col = month_cols.get('Actual')
            accrual_col = month_cols.get('Accrual')
            if actual_col and accrual_col:
                actual_idx = column_index_from_string(actual_col)
                variance_col_letter = get_column_letter(actual_idx + 1)
                
                # Check if this column represents the monthly variance column
                header_val = self.sheet.cell(row=self.header_row, column=actual_idx + 1).value
                if header_val and 'variance' in str(header_val).lower():
                    formula = f'=IF(OR({accrual_col}{ppi_row}="",{actual_col}{ppi_row}=""),"",{accrual_col}{ppi_row}-{actual_col}{ppi_row})'
                    self.sheet.cell(row=ppi_row, column=actual_idx + 1, value=formula)
                    
        # 2. Update/Rewrite other existing formulas (e.g., BR Column Total in projects template)
        # to use the new ppi_row index.
        import re
        for col_idx in range(2, (self.sheet.max_column or 200) + 1):
            cell = self.sheet.cell(row=ppi_row, column=col_idx)
            v = cell.value
            if v is not None and isinstance(v, str) and v.startswith('='):
                m = re.search(r'[A-Z]+\$?(\d+)', v)
                if m:
                    old_row = m.group(1)
                    if str(old_row) != str(ppi_row):
                        pattern = r'([A-Z]+\$?)' + str(old_row)
                        new_formula = re.sub(pattern, r'\g<1>' + str(ppi_row), v)
                        if new_formula != v:
                            cell.value = new_formula

        # 3. Dynamic Grand Total formula for Previous Period Invoices row
        if self.total_col:
            self._write_total_formula(ppi_row)

    # Map full/variant month words found in header cells → canonical 3-letter key
    _HEADER_MONTH_ALIASES = {
        "jan": "Jan", "feb": "Feb",
        "mar": "Mar", "march": "Mar",
        "apr": "Apr", "april": "Apr",
        "may": "May",
        "jun": "Jun", "june": "Jun",
        "jul": "Jul", "july": "Jul",
        "aug": "Aug",
        "sep": "Sep", "sept": "Sep",
        "oct": "Oct", "nov": "Nov", "dec": "Dec",
    }

    def _find_actual_header_row(self) -> int:
        """Scan every cell in every row to find the real header row.

        Two strategies are tried in order:

        1. Look for a cell containing 'contact for po' (any column) — the
           explicit label used in both the OpEx and project templates.
        2. Look for a row that contains at least two metric+month header cells
           (e.g. 'Forecast Jan', 'Actual Feb') — the same tokens that
           get_column_map parses.  This catches templates where the label cell
           is absent or worded differently.

        Falls back to self.header_row (the config value) when neither signal
        is found, so blank/unfamiliar templates still work.
        """
        _metric_kws = ("forecast", "accrual reversal", "accrual", "actual")
        _month_kws  = set(self._HEADER_MONTH_ALIASES.keys())
        max_col = self.sheet.max_column or 50
        max_row = self.sheet.max_row or 1000

        for r in range(1, max_row + 1):
            metric_month_hits = 0
            for c in range(1, max_col + 1):
                v = self.sheet.cell(row=r, column=c).value
                if v is None:
                    continue
                text = str(v).strip().lower()
                # Strategy 1 — explicit label
                if "contact for po" in text:
                    return r
                # Strategy 2 — metric+month header cell
                for kw in _metric_kws:
                    if text.startswith(kw):
                        words = text.split()
                        if words and words[-1].rstrip('.,') in _month_kws:
                            metric_month_hits += 1
                        break
            if metric_month_hits >= 2:
                return r

        return self.header_row  # fallback

    def get_column_map(self, starting_col):
        """Build month→metric→column-letter map by scanning the template header row.

        The header row contains cells like 'Forecast Jan', 'Accrual Reversal Dec',
        'Actual   Feb', etc.  We parse each cell to extract the metric keyword and
        the month keyword, normalise both, and record the column letter.

        The first 'Accrual Reversal Dec' column (col N in the current template) is
        the prior-year December accrual reversal — mapped to the key 'Dec (PY)' so
        it stays distinct from the current-year 'Dec' block.

        Falls back to the old fixed-offset arithmetic if no header cells are found
        (e.g. a blank template that hasn't been opened in Excel yet).
        """
        metrics_lower = {
            "accrual reversal": "Accrual Reversal",
            "forecast":         "Forecast",
            "accrual":          "Accrual",
            "actual":           "Actual",
        }

        col_map: dict[str, dict[str, str]] = {}
        dec_py_seen = False  # first 'Dec' Accrual Reversal is Dec (PY)
        max_col = self.sheet.max_column or 200

        # Use the actual header row position rather than the config value so this
        # works correctly on blank templates and row-shifted files.
        actual_header_row = self._find_actual_header_row()

        for col_idx in range(1, max_col + 1):
            raw = self.sheet.cell(row=actual_header_row, column=col_idx).value
            if not raw:
                continue
            text = str(raw).strip()

            # Identify which metric this cell represents
            matched_metric = None
            for kw, canonical in metrics_lower.items():
                if text.lower().startswith(kw):
                    matched_metric = canonical
                    break
            if matched_metric is None:
                continue

            # Extract the month word (last token in the cell text)
            words = text.split()
            month_word = words[-1].lower().rstrip('.,').strip() if words else ""
            month_key = self._HEADER_MONTH_ALIASES.get(month_word)
            if month_key is None:
                continue

            # The very first 'Accrual Reversal Dec' is the prior-year Dec column
            if matched_metric == "Accrual Reversal" and month_key == "Dec" and not dec_py_seen:
                month_key = "Dec (PY)"
                dec_py_seen = True

            if month_key not in col_map:
                col_map[month_key] = {}

            # Guard against duplicate headers (e.g. two "Accrual May" cells where
            # the second should have been "Actual May").  If this metric slot is
            # already filled for this month, check whether the *next* unused metric
            # for this month fits better and warn so the template can be corrected.
            if matched_metric in col_map[month_key]:
                existing_col = col_map[month_key][matched_metric]
                # Determine which metric slot is missing for this month and fill it
                expected_order = ["Forecast", "Accrual", "Actual", "Accrual Reversal"]
                missing = [m for m in expected_order if m not in col_map[month_key]]
                if missing:
                    fallback = missing[0]
                    print(
                        f"WARNING: Duplicate header '{text}' at col "
                        f"{get_column_letter(col_idx)} (already mapped from col "
                        f"{existing_col}). Treating as '{fallback} {month_key}' — "
                        f"please correct the template header."
                    )
                    col_map[month_key][fallback] = get_column_letter(col_idx)
                else:
                    print(
                        f"WARNING: Duplicate header '{text}' at col "
                        f"{get_column_letter(col_idx)} ignored — all slots for "
                        f"{month_key} are already filled."
                    )
                continue

            col_map[month_key][matched_metric] = get_column_letter(col_idx)

        if col_map:
            return col_map

        # ----------------------------------------------------------------
        # Fallback: old fixed-offset arithmetic (used when the template has
        # no populated header row, e.g. a truly blank workbook).
        # ----------------------------------------------------------------
        print(
            "WARNING: No metric headers found in header row — "
            "falling back to fixed column offsets starting at column "
            f"'{starting_col}'."
        )
        months_fb = [
            "Dec (PY)",
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
        ]
        metrics_fb = ["Accrual Reversal", "Forecast", "Accrual", "Actual"]
        col_index = column_index_from_string(starting_col)
        col_map_fb: dict[str, dict[str, str]] = {}
        for month in months_fb:
            month_map: dict[str, str] = {}
            for metric in metrics_fb:
                month_map[metric] = get_column_letter(col_index)
                col_index += 1
            col_map_fb[month] = month_map
            col_index += 1  # skip variance column
        return col_map_fb

    def insert_missing_po_rows(self, hierarchy: dict, pos: dict[str, int]) -> dict[str, int]:
        """
        For each cost center in the hierarchy:
        - If the cost center already has PO rows in the template, insert only the
          rows that are missing (including RECLASS) directly after its last existing row.
        - If the cost center has no rows at all, insert all PO rows (including RECLASS)
          just above the 'Previous Period Invoices' stop marker.
        Existing written rows are never modified.
        """
        stop_row = None
        for row_idx in range(1, (self.sheet.max_row or 1000) + 1):
            cell_val = self.sheet[f"A{row_idx}"].value
            if cell_val is not None and str(cell_val).strip() == "Previous Period Invoices":
                stop_row = row_idx
                break

        if stop_row is None:
            print("WARNING: 'Previous Period Invoices' marker not found – PO rows not inserted.")
            return pos

        po_col_idx = column_index_from_string(self.po_column)
        # Use configured p3_id_column when available (project template), otherwise
        # fall back to column J (index 10) used by the OpEx template.
        cc_col_idx = self.p3_id_column if self.p3_id_column else 10
        wbs_col_idx = 6  # Column F
        max_col = self.sheet.max_column or 1

        # Determine whether a dedicated CC / P3-ID label column exists in the data rows.
        # If cc_col_idx is occupied by a metric column (e.g. "Accrual Reversal Dec" in the
        # project template) we must NOT write the label there.
        _data_col_indices: set[int] = set()
        for _mc in self.column_map.values():
            for _cl in _mc.values():
                _data_col_indices.add(column_index_from_string(_cl))
        _has_cc_label_col: bool = (
            False if self.p3_id_column is not None
            else (
                cc_col_idx not in _data_col_indices          # col J is free (OpEx template)
                and cc_col_idx != self.po_status_col         # never write P3 ID into PO Status
            )
        )

        # Build a reverse map: po_number → cc_id from the hierarchy so that
        # _scan_existing_rows_by_cc can find which group owns each existing PO
        # row even when there is no explicit CC-label column in the sheet
        # (e.g. the project template).
        _po_to_cc: dict[str, str] = {}
        for _cc_id, _cc_obj in hierarchy.items():
            for _wbs_code, _wbs_obj in _cc_obj.wbs_codes.items():
                for _po_num in _wbs_obj.pos:
                    _po_to_cc[self._norm_po(_po_num)] = _cc_id

        def _get_ref_styles(ref_row: int) -> tuple[dict, float | None]:
            """Snapshot per-column styles from ref_row."""
            styles = {}
            for col_idx in range(1, max_col + 1):
                src = self.sheet.cell(row=ref_row, column=col_idx)
                styles[col_idx] = {
                    'font': copy(src.font),
                    'border': copy(src.border),
                    'alignment': copy(src.alignment),
                    'number_format': src.number_format,
                    'fill': _copy_fill(src.fill),
                    'comment': copy(src.comment),
                }
            return styles, self.sheet.row_dimensions[ref_row].height

        def _scan_existing_rows_by_cc() -> dict:
            """Re-scan the sheet to get the last existing row per cost center / P3 ID.

            Two strategies are tried in order:
            1. If a dedicated CC/P3-ID label column exists (cc_col_idx has a value
               and it isn't a known data column), read the label directly from that cell.
            2. Otherwise resolve the CC for each row via the PO→CC reverse map built
               from the hierarchy (used by the project template where there is no
               separate label column in the data rows).
            """
            result = {}
            for row_idx in range(self.header_row + 1, (self.sheet.max_row or 1000) + 1):
                cell_val = self.sheet[f"A{row_idx}"].value
                if cell_val is not None and str(cell_val).strip() == "Previous Period Invoices":
                    break
                po_val = self.sheet.cell(row=row_idx, column=po_col_idx).value
                po_text = self._norm_po(str(po_val).strip()) if po_val is not None and str(po_val).strip() else None
                if not po_text:
                    continue
                # Strategy 1: read CC label directly from the dedicated column
                cc_text = None
                if self.p3_id_column:
                    cc_raw = self.sheet.cell(row=row_idx, column=cc_col_idx).value
                    if cc_raw is not None and str(cc_raw).strip():
                        cc_text = str(cc_raw).strip().split('/')[0].strip()
                # Strategy 2: look up the PO in the hierarchy reverse map
                if not cc_text:
                    cc_text = _po_to_cc.get(po_text)
                if cc_text:
                    result[cc_text] = row_idx  # keep updating → ends at last row for this CC
            return result

        inserted = []
        for cc_id, cost_center in hierarchy.items():
            # Collect (po_number, wbs_code) pairs for POs not yet in the template
            po_numbers = []
            po_wbs_map = {}      # po_number -> wbs_code
            po_le_map = {}       # po_number -> legal_entity
            po_country_map = {}  # po_number -> country
            po_vendor_map = {}   # po_number -> vendor_name
            po_gl_map = {}       # po_number -> gl_account
            po_gross_map = {}    # po_number -> gross_po_value
            po_req_map = {}      # po_number -> req_title
            for wbs_code, wbs in cost_center.wbs_codes.items():
                if wbs_code.upper() == "ER":
                    continue
                for po_number, po_obj in wbs.pos.items():
                    norm_po = self._norm_po(po_number)
                    if norm_po not in pos and norm_po not in po_wbs_map:
                        po_number = norm_po
                        po_numbers.append(po_number)
                        po_wbs_map[po_number] = wbs_code
                        po_le_map[po_number] = po_obj.legal_entity
                        po_country_map[po_number] = po_obj.country
                        po_vendor_map[po_number] = po_obj.vendor_name
                        po_gl_map[po_number] = po_obj.gl_account
                        po_gross_map[po_number] = po_obj.gross_po_value
                        po_req_map[po_number] = po_obj.req_title

            if not po_numbers:
                continue

            # Re-scan each time so row numbers stay correct after previous insertions
            existing_last_row_by_cc = _scan_existing_rows_by_cc()

            if cc_id in existing_last_row_by_cc:
                # CC already has rows — insert new rows immediately after the last one
                insert_at = existing_last_row_by_cc[cc_id] + 1
            else:
                # CC has no rows yet — find the stop marker then scan backward
                # past any trailing blank rows so inserted rows land directly
                # after the last content row (no blank gap).
                stop_row_cur = None
                for row_idx in range(1, (self.sheet.max_row or 1000) + 1):
                    if self.sheet[f"A{row_idx}"].value is not None and \
                            str(self.sheet[f"A{row_idx}"].value).strip() == "Previous Period Invoices":
                        stop_row_cur = row_idx
                        break
                if stop_row_cur is None:
                    print("WARNING: stop marker disappeared during insertion.")
                    break
                # Walk backward from stop_row_cur - 1 looking for the last row
                # that has a real PO number in the PO column.  We intentionally
                # ignore everything else (labels, zero-filled variance formulas,
                # cost-center labels, etc.) because those are template scaffolding
                # that appears in placeholder rows even when no PO has been written.
                # If no PO row is found at all, fall back to header_row + 1 so the
                # first insertion lands directly under the header.
                insert_at = self.header_row + 1
                for row_idx in range(stop_row_cur - 1, self.header_row, -1):
                    po_val = self.sheet.cell(row=row_idx, column=po_col_idx).value
                    if po_val is not None:
                        po_str = str(po_val).strip()
                        if po_str and po_str.lower() != "none" and not po_str.startswith("="):
                            insert_at = row_idx + 1
                            break

            for po_number in po_numbers:
                # Use the row immediately before the insertion point (but not the
                # header row) as the style reference, so inserted rows inherit the
                # same fill and font as their adjacent neighbours regardless of
                # whether those neighbours have a PO number in them.
                style_source_row = self.header_row + 1   # safe fallback
                for _sr in range(insert_at - 1, self.header_row, -1):
                    # Skip the header row itself (bold/coloured) — use any other row
                    if _sr != self.header_row:
                        style_source_row = _sr
                        break
                ref_styles, ref_height = _get_ref_styles(style_source_row)
                self.sheet.insert_rows(insert_at)
                for col_idx in range(1, max_col + 1):
                    new_cell = self.sheet.cell(row=insert_at, column=col_idx)
                    # openpyxl copies the value from the displaced row into the
                    # new row — clear it so no stale formula survives.
                    new_cell.value = None
                    s = ref_styles[col_idx]
                    new_cell.font = copy(s['font'])
                    new_cell.border = copy(s['border'])
                    new_cell.alignment = copy(s['alignment'])
                    new_cell.number_format = s['number_format']
                    new_cell.fill = _copy_fill(s['fill'])
                    new_cell.comment = copy(s['comment'])
                if ref_height:
                    self.sheet.row_dimensions[insert_at].height = ref_height

                vendor_val = po_vendor_map.get(po_number)
                if vendor_val is not None:
                    self.sheet.cell(row=insert_at, column=5, value=vendor_val)
                # Write the CC / P3-ID label only when a dedicated label column exists
                # (controlled by _has_cc_label_col set above the loop).
                if _has_cc_label_col:
                    self.sheet.cell(row=insert_at, column=cc_col_idx, value=cc_id)
                self.sheet.cell(row=insert_at, column=po_col_idx, value=po_number)
                self.sheet.cell(row=insert_at, column=wbs_col_idx, value=po_wbs_map.get(po_number))
                if self.p3_id_column is None:  # OpEx only
                    gl_val = po_gl_map.get(po_number)
                    if gl_val is not None:
                        self.sheet.cell(row=insert_at, column=7, value=gl_val)
                    le_val = po_le_map.get(po_number)
                    if le_val is not None:
                        self.sheet.cell(row=insert_at, column=9, value=le_val)
                    country_val = po_country_map.get(po_number)
                    if country_val is not None:
                        self.sheet.cell(row=insert_at, column=10, value=country_val)
                gross_val = po_gross_map.get(po_number)
                if gross_val is not None and self.po_value_col is not None and self.p3_id_column is None:
                    self.sheet.cell(row=insert_at, column=self.po_value_col, value=round(gross_val, 2))
                req_val = po_req_map.get(po_number)
                if req_val is not None:
                    for _col in self.req_title_cols:
                        self.sheet.cell(row=insert_at, column=_col, value=req_val)
                pos[po_number] = insert_at
                inserted.append(po_number)

                # Every insertion shifts all rows below by 1 — update pos accordingly
                for key in pos:
                    if key != po_number and pos[key] >= insert_at:
                        pos[key] += 1

                insert_at += 1

        if inserted:
            print(f"Inserted {len(inserted)} PO row(s): {inserted}")
        else:
            print("No additional PO rows needed.")
        return pos

    def insert_er_rows(self, hierarchy: dict, pos: dict[str, int]) -> dict[str, int]:
        """
        Collects all ER numbers from the hierarchy (those assigned WBS="ER" during
        exception processing) and inserts a new row for each one in the Template sheet
        directly above the 'Previous Period Invoices' stop-marker row.

        The new rows have only the ER number written into the PO column (column B by
        default, matching self.po_column).  All other cells are left blank so the
        normal write_hierarchy call can fill in the monthly data.

        Args:
            hierarchy: The built hierarchy dict returned by build_hierarchy.
            pos: The existing PO → row mapping from TemplateReader.get_existing_pos().

        Returns:
            Updated pos dict that includes the newly inserted ER rows.
        """
        # Collect unique ER numbers that are NOT already in the template
        er_pattern = re.compile(r'^ER\d+$', re.IGNORECASE)
        er_numbers = []
        er_cc_map = {}       # er_number -> cost_center_id
        er_le_map = {}       # er_number -> legal_entity
        er_country_map = {}  # er_number -> country
        er_vendor_map = {}   # er_number -> vendor_name
        er_gl_map = {}       # er_number -> gl_account
        er_gross_map = {}    # er_number -> gross_po_value
        er_req_map = {}      # er_number -> req_title
        er_wbs_map = {}      # er_number -> real WBS (po.real_wbs if set, else "ER")
        seen = set()
        for cc_id, cost_center in hierarchy.items():
            for wbs_code, wbs in cost_center.wbs_codes.items():
                if wbs_code.upper() == "ER":
                    for po_number, po_obj in wbs.pos.items():
                        if er_pattern.match(po_number) and po_number not in pos and po_number not in seen:
                            er_numbers.append(po_number)
                            er_cc_map[po_number] = cc_id
                            er_le_map[po_number] = po_obj.legal_entity
                            er_country_map[po_number] = po_obj.country
                            er_vendor_map[po_number] = po_obj.vendor_name
                            er_gl_map[po_number] = po_obj.gl_account
                            er_gross_map[po_number] = po_obj.gross_po_value
                            er_req_map[po_number] = po_obj.req_title
                            er_wbs_map[po_number] = po_obj.real_wbs or "ER"
                            seen.add(po_number)

        if not er_numbers:
            print("No ER rows to insert into template.")
            return pos

        # Find the stop-marker row (Previous Period Invoices) in the sheet
        stop_row = None
        for row_idx in range(1, (self.sheet.max_row or 1000) + 1):
            cell_val = self.sheet[f"A{row_idx}"].value
            if cell_val is not None and str(cell_val).strip() == "Previous Period Invoices":
                stop_row = row_idx
                break

        if stop_row is None:
            print("WARNING: 'Previous Period Invoices' marker not found – ER rows not inserted.")
            return pos

        # Update header label in the PO column to reflect ER numbers
        po_col_idx = column_index_from_string(self.po_column)
        header_cell = self.sheet.cell(row=self.header_row, column=po_col_idx)
        header_cell.value = "PO Number/ER Number"

        # Light green fill for ER cells
        green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")

        max_col = self.sheet.max_column or 1

        # Reuse the same CC-label-column logic as insert_missing_po_rows.
        _er_cc_col_idx = self.p3_id_column if self.p3_id_column else 10
        _er_data_cols: set[int] = set()
        for _mc in self.column_map.values():
            for _cl in _mc.values():
                _er_data_cols.add(column_index_from_string(_cl))
        _er_has_cc_col: bool = (
            False if self.p3_id_column is not None
            else (
                _er_cc_col_idx not in _er_data_cols          # col J is free (OpEx template)
                and _er_cc_col_idx != self.po_status_col     # never write P3 ID into PO Status
            )
        )

        def _get_er_ref_styles(ref_row: int) -> tuple[dict, float | None]:
            """Snapshot per-column styles from ref_row for ER insertion."""
            styles = {}
            for col_idx in range(1, max_col + 1):
                src = self.sheet.cell(row=ref_row, column=col_idx)
                styles[col_idx] = {
                    'font': copy(src.font),
                    'border': copy(src.border),
                    'alignment': copy(src.alignment),
                    'number_format': src.number_format,
                    'fill': _copy_fill(src.fill),
                    'comment': copy(src.comment),
                }
            return styles, self.sheet.row_dimensions[ref_row].height

        # Find the last row that has a real PO number and insert ER rows after it.
        # Falls back to header_row + 1 when no PO rows exist yet.
        insert_at = self.header_row + 1
        for row_idx in range(stop_row - 1, self.header_row, -1):
            po_val = self.sheet.cell(row=row_idx, column=po_col_idx).value
            if po_val is not None:
                po_str = str(po_val).strip()
                if po_str and po_str.lower() != "none" and not po_str.startswith("="):
                    insert_at = row_idx + 1
                    break

        for er in er_numbers:
            # Use the row immediately before the insertion point (not the header)
            # as the style reference so inserted rows match their neighbours.
            style_source_row = self.header_row + 1   # safe fallback
            for _sr in range(insert_at - 1, self.header_row, -1):
                if _sr != self.header_row:
                    style_source_row = _sr
                    break
            ref_styles, ref_height = _get_er_ref_styles(style_source_row)
            self.sheet.insert_rows(insert_at)

            # Apply reference row formatting to every cell in the new row
            for col_idx in range(1, max_col + 1):
                new_cell = self.sheet.cell(row=insert_at, column=col_idx)
                # Clear any value openpyxl copied from the displaced row.
                new_cell.value = None
                s = ref_styles[col_idx]
                new_cell.font = copy(s['font'])
                new_cell.comment = copy(s['comment'])
                # Normalise top border: always use 'thin' so each new row looks
                # identical to the reference data rows regardless of insertion order
                orig_border = s['border']
                new_cell.border = Border(
                    left=copy(orig_border.left),
                    right=copy(orig_border.right),
                    top=Side(border_style='thin'),
                    bottom=copy(orig_border.bottom),
                )
                new_cell.alignment = copy(s['alignment'])
                new_cell.number_format = s['number_format']
                # Fill: green on PO column, same as neighbour on all other columns
                if col_idx == po_col_idx:
                    new_cell.fill = green_fill
                else:
                    new_cell.fill = _copy_fill(s['fill'])

            if ref_height:
                self.sheet.row_dimensions[insert_at].height = ref_height

            # Write ER number into the PO column and WBS into column F
            self.sheet.cell(row=insert_at, column=po_col_idx, value=er)
            self.sheet.cell(row=insert_at, column=6, value=er_wbs_map.get(er, "ER"))
            # Write Vendor Name into column E
            vendor_val = er_vendor_map.get(er)
            if vendor_val is not None:
                self.sheet.cell(row=insert_at, column=5, value=vendor_val)
            # Write Cost Center / P3-ID label only when a dedicated column exists.
            cc_val = er_cc_map.get(er)
            if cc_val is not None and _er_has_cc_col:
                self.sheet.cell(row=insert_at, column=_er_cc_col_idx, value=cc_val)
            if self.p3_id_column is None:  # OpEx only
                # Write G/L Account into column G
                gl_val = er_gl_map.get(er)
                if gl_val is not None:
                    self.sheet.cell(row=insert_at, column=7, value=gl_val)
                # Write LE into column I
                le_val = er_le_map.get(er)
                if le_val is not None:
                    self.sheet.cell(row=insert_at, column=9, value=le_val)
                # Write Country into column J
                country_val = er_country_map.get(er)
                if country_val is not None:
                    self.sheet.cell(row=insert_at, column=10, value=country_val)
            # Write PO total into the correct column (Gross PO Value or Invoice Amount)
            gross_val = er_gross_map.get(er)
            if gross_val is not None and self.po_value_col is not None and self.p3_id_column is None:
                self.sheet.cell(row=insert_at, column=self.po_value_col, value=round(gross_val, 2))
            # Write Requisition Title into all matching columns
            req_val = er_req_map.get(er)
            if req_val is not None:
                for _col in self.req_title_cols:
                    self.sheet.cell(row=insert_at, column=_col, value=req_val)

            pos[er] = insert_at
            insert_at += 1  # next ER goes after the one just inserted

        print(f"Inserted {len(er_numbers)} ER row(s) into template: {er_numbers}")
        return pos

    def _get_comments_col(self) -> str | None:
        """Scan the header row for a 'Comments' column. Returns the column letter or None."""
        for col in range(1, (self.sheet.max_column or 200) + 1):
            val = self.sheet.cell(row=self.header_row, column=col).value
            if val and 'comment' in str(val).strip().lower():
                return get_column_letter(col)
        return None

    def write_hierarchy(self, hierarchy: dict, pos: dict[str, int]):
        '''
        Writes hierarchy to template.
        Iterates CostCenter -> WBSCode -> PO and writes MonthlyMetrics to correct cells.
        Only writes to blank cells unless overwrite=True.
        '''
        reclass_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        comments_col = self._get_comments_col()

        # Project template: wipe the entire PO Total column across all data rows so
        # no pre-existing formula or value survives regardless of whether the PO is
        # in the hierarchy.
        if self.p3_id_column is not None and self.po_value_col is not None:
            for _r in range(self.header_row + 1, (self.sheet.max_row or 1000) + 1):
                _sentinel = self.sheet.cell(row=_r, column=1).value
                if _sentinel is not None and str(_sentinel).strip() == "Previous Period Invoices":
                    break
                self.sheet.cell(row=_r, column=self.po_value_col).value = None

        # Determine the CC/P3-ID label column, and whether one exists at all.
        _cc_write_col = self.p3_id_column if self.p3_id_column else 10
        _wh_data_cols: set[int] = set()
        for _mc in self.column_map.values():
            for _cl in _mc.values():
                _wh_data_cols.add(column_index_from_string(_cl))
        _has_cc_label_col: bool = (
            False if self.p3_id_column is not None
            else (
                _cc_write_col not in _wh_data_cols           # col J is free (OpEx template)
                and _cc_write_col != self.po_status_col      # never write P3 ID into PO Status
            )
        )

        for cc_id, cost_center in hierarchy.items():
            for wbs_code, wbs in cost_center.wbs_codes.items():
                for po_number, po in wbs.pos.items():
                    # Normalize PO key before lookup so "9500905777.0" matches "9500905777"
                    po_number = self._norm_po(po_number)
                    # Skip if PO not in template
                    if po_number not in pos:
                        print(f"PO '{po_number}' not found in template. Skipping.")
                        continue
                    row = pos[po_number]

                    # Cost Center / P3-ID label (skipped for project template — no label column)
                    if _has_cc_label_col:
                        cc_cell = self.sheet.cell(row=row, column=_cc_write_col)
                        if self._should_write(cc_cell.value):
                            cc_cell.value = cc_id

                    # Vendor name (col E) — also skip "not assigned" placeholder
                    if po.vendor_name is not None:
                        vendor_cell = self.sheet.cell(row=row, column=5)
                        existing = vendor_cell.value
                        if not (not self.overwrite and isinstance(existing, str) and existing.startswith('=')):
                            existing_str = str(existing).strip() if existing is not None else ""
                            if self.overwrite or existing_str == "" or existing_str.lower() == "not assigned":
                                vendor_cell.value = po.vendor_name

                    # WBS code (col F) — use real_wbs for ER rows when available
                    wbs_to_write = po.real_wbs if (wbs_code == "ER" and po.real_wbs) else wbs_code
                    wbs_cell = self.sheet.cell(row=row, column=6)
                    if self._should_write(wbs_cell.value):
                        wbs_cell.value = wbs_to_write

                    if self.p3_id_column is None:  # OpEx only
                        # G/L Account (col G)
                        if po.gl_account is not None:
                            gl_cell = self.sheet.cell(row=row, column=7)
                            if self._should_write(gl_cell.value):
                                gl_cell.value = po.gl_account

                        # Legal Entity (col I)
                        if po.legal_entity is not None:
                            le_cell = self.sheet.cell(row=row, column=9)
                            if self._should_write(le_cell.value):
                                le_cell.value = po.legal_entity

                        # Country (col J)
                        if po.country is not None:
                            country_cell = self.sheet.cell(row=row, column=10)
                            if self._should_write(country_cell.value):
                                country_cell.value = po.country

                    # PO total (Gross PO Value or Invoice Amount column)
                    if self.po_value_col is not None:
                        gross_cell = self.sheet.cell(row=row, column=self.po_value_col)
                        if self.p3_id_column is not None:
                            # Project template — always wipe whatever is there (formula or value).
                            gross_cell.value = None
                        elif po.gross_po_value is not None and self._should_write(gross_cell.value):
                            gross_cell.value = round(po.gross_po_value, 2)

                    # Requisition title
                    if po.req_title is not None:
                        for _col in self.req_title_cols:
                            req_cell = self.sheet.cell(row=row, column=_col)
                            if self._should_write(req_cell.value):
                                req_cell.value = po.req_title

                    # Project Name (projects only) — keep blank
                    if self.p3_id_column is not None and self.project_name_col is not None:
                        self.sheet.cell(row=row, column=self.project_name_col).value = None

                    # Monthly metric values
                    for month, metrics in po.monthly_data.items():
                        if month not in self.column_map:
                            continue
                        month_cols = self.column_map[month]
                        values = {
                            'Accrual Reversal': metrics.accrual_reversal,
                            'Forecast':         metrics.forecast,
                            'Accrual':          metrics.accrual,
                            'Actual':           metrics.actual,
                        }
                        for metric, col_letter in month_cols.items():
                            cell = self.sheet[f"{col_letter}{row}"]
                            value = values.get(metric)
                            # Skip zero/None to avoid polluting blank cells with 0.
                            if not self.overwrite and (value is None or value == 0):
                                continue
                            if self._should_write(cell.value):
                                cell.value = value

                        actual_col  = month_cols.get('Actual')
                        accrual_col = month_cols.get('Accrual')
                        has_actual  = metrics.actual  is not None and metrics.actual  != 0
                        has_accrual = metrics.accrual is not None and metrics.accrual != 0
                        if actual_col and accrual_col and (has_actual or has_accrual):
                            variance_col  = get_column_letter(column_index_from_string(actual_col) + 1)
                            variance_cell = self.sheet[f"{variance_col}{row}"]
                            if self._should_write(variance_cell.value):
                                variance_cell.value = f"={accrual_col}{row}-{actual_col}{row}"

                        # Highlight and annotate Actual cell if reclass adjustments exist
                        reclass_entries = po.reclass_adjustments.get(month, [])
                        if reclass_entries and actual_col:
                            actual_cell = self.sheet[f"{actual_col}{row}"]
                            actual_cell.fill = reclass_fill
                            total_reclass = sum(amt for amt, _ in reclass_entries)
                            original = (metrics.actual or 0) - total_reclass
                            lines = [
                                f"Original Amount: ${original:,.2f}",
                                "Reclass adjustment(s) included in Actual:",
                            ]
                            for amt, desc in reclass_entries:
                                sign = "+" if amt >= 0 else ""
                                lines.append(f"  {sign}${amt:,.2f}")
                                if desc:
                                    lines.append(f"  Description: {desc}")
                            lines.append(f"Adjusted Total: ${metrics.actual:,.2f}")
                            
                            actual_cell.comment = _append_comment(
                                actual_cell.comment,
                                "\n".join(lines),
                            )

                    # Total formula — only when the PO has monthly data
                    if po.monthly_data:
                        self._write_total_formula(row)

        # After all PO data is written, refresh the summary row formulas so they
        # cover the full data range (rows header_row+1 … last data row).
        self._update_summary_formulas()


    def write_forecast_source_sheet(self, forecast_df, pos: dict[str, int]):
        """Write forecast data to a new 'Forecast Source Data' sheet, filtered to template POs."""
        # Filter to template POs
        if self.forecast_po_col not in forecast_df.columns:
            raise KeyError(
                f"The expected PO column '{self.forecast_po_col}' was not found in the forecast dataset. "
                f"Please verify the column mappings or configuration in config_project.yaml / config_base.yaml. "
                f"Available columns in the loaded forecast file: {list(forecast_df.columns)}"
            )

        forecast_df[self.forecast_po_col] = (
            forecast_df[self.forecast_po_col]
            .apply(lambda x: str(int(float(x))) if str(x).replace('.','',1).isdigit() else str(x))
        )
        filtered_df = forecast_df[forecast_df[self.forecast_po_col].isin(pos.keys())]

        visible_cols = [c for c in self.forecast_source_cols if c in filtered_df.columns]
        hidden_cols = [c for c in filtered_df.columns if c not in visible_cols]

        ordered_cols = visible_cols + hidden_cols
        source_df = filtered_df[ordered_cols]

        ws = self.wb.create_sheet("Forecast Source Data")

        # Styled header matching the Exceptions tab format
        hdr_fill = PatternFill("solid", fgColor="1F4E79")   # Dark blue header
        hdr_font = Font(bold=True, color="FFFFFF")          # Bold white font

        for col_idx, col_name in enumerate(source_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = hdr_font
            cell.fill = hdr_fill

        for row_idx, (_, row) in enumerate(source_df.iterrows(), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        data_start = 2
        data_end = ws.max_row
        total_row = data_end + 1
        ws.cell(row=total_row, column=1, value="PO Total")

        for col_idx, col_name in enumerate(source_df.columns, start=1):
            if col_name in visible_cols and col_name != self.forecast_po_col:
                letter = get_column_letter(col_idx)
                formula = f"=SUBTOTAL(9,{letter}{data_start}:{letter}{data_end})"
                ws.cell(row=total_row, column=col_idx, value=formula)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        for col_idx, col_name in enumerate(source_df.columns, start=1):
            letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for row_idx in range(2, len(source_df) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = max_len + 2

        if hidden_cols:
            start_idx = len(visible_cols) + 1
            end_idx = len(visible_cols) + len(hidden_cols)
            ws.column_dimensions.group(
                get_column_letter(start_idx),
                get_column_letter(end_idx),
                hidden=True
            )

    def write_transactional_source_sheet(self, transactions_df, pos: dict[str, int]):
        """Write transactional detail to a new 'Transactions Source Data' sheet.
        Includes rows where PO is in the template, plus all Reclass rows."""
        if self.transactional_po_col not in transactions_df.columns:
            raise KeyError(
                f"The expected PO column '{self.transactional_po_col}' was not found in the transactional dataset. "
                f"Please verify your file layout or configuration mappings in config_project.yaml / config_base.yaml. "
                f"Available columns in the loaded transactional file: {list(transactions_df.columns)}"
            )

        transactions_df[self.transactional_po_col] = transactions_df[self.transactional_po_col].astype(str)
        in_template = transactions_df[self.transactional_po_col].isin(pos.keys())
        is_reclass = transactions_df['Type'] == 'Reclass' if 'Type' in transactions_df.columns else False
        source_df = transactions_df[in_template | is_reclass]

        visible_cols = [c for c in self.transactional_source_cols if c in source_df.columns]
        hidden_cols = [c for c in source_df.columns if c not in visible_cols]

        final_cols = visible_cols + hidden_cols
        source_df = source_df[final_cols]

        ws = self.wb.create_sheet("Transactions Source Data")

        # Determine if we are running the projects pipeline
        is_project = self.p3_id_column is not None

        # Headers mapping for project visual consistency
        header_remap = {}
        if is_project:
            header_remap = {
                'Accounting Period': 'Fiscal Year/Period',
                'AP Voucher Number': 'Vendor Invoice',
                'PO Number':         'Document num/PO#',
            }

        # Styled header matching the Exceptions tab format
        hdr_fill = PatternFill("solid", fgColor="1F4E79")   # Dark blue header
        hdr_font = Font(bold=True, color="FFFFFF")          # Bold white font

        for col_idx, col_name in enumerate(source_df.columns, start=1):
            display_name = header_remap.get(col_name, col_name)
            cell = ws.cell(row=1, column=col_idx, value=display_name)
            cell.font = hdr_font
            cell.fill = hdr_fill

        for row_idx, (_, row) in enumerate(source_df.iterrows(), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        if not is_project and len(source_df) > 0:
            data_start = 2
            data_end = ws.max_row
            total_row = data_end + 1
            ws.cell(row=total_row, column=1, value="PO Total")

            for col_idx, col_name in enumerate(source_df.columns, start=1):
                if col_name in visible_cols and col_name != self.transactional_po_col:
                    letter = get_column_letter(col_idx)
                    formula = f"=SUBTOTAL(9,{letter}{data_start}:{letter}{data_end})"
                    ws.cell(row=total_row, column=col_idx, value=formula)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        for col_idx, col_name in enumerate(source_df.columns, start=1):
            letter = get_column_letter(col_idx)
            display_name = header_remap.get(col_name, col_name)
            max_len = len(str(display_name))
            for row_idx in range(2, len(source_df) + 2):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[letter].width = min(max_len + 2, 55)

        if hidden_cols:
            start_idx = len(visible_cols) + 1
            end_idx = len(visible_cols) + len(hidden_cols)
            ws.column_dimensions.group(
                get_column_letter(start_idx),
                get_column_letter(end_idx),
                hidden=True
            )


    
    @staticmethod
    def _extract_er_number(po_value):
        """If po_value contains an ER number (e.g. 'ER97054 - ARCH & TECH - ...'), return only 'ER97054'.
        Otherwise return the original value unchanged."""
        if po_value and isinstance(po_value, str):
            match = re.match(r'^(ER\d+)', po_value.strip(), re.IGNORECASE)
            if match:
                return match.group(1).upper()
        return po_value

    def write_exception_sheet(self, exception_log, transactional_df, pos: dict | None = None):
        """Write the Exceptions sheet.

        Shows all exceptions in a single unified flat table starting at Row 1,
        with auto-filter across all columns, freeze panes at Row 2, and auto-fitted
        column widths, exactly matching the formatting of the other source tabs.
        """
        ws = self.wb.create_sheet("Exceptions")

        # Compile all active exceptions
        active_entries = []
        for e in exception_log.entries:
            if e.exception_type == ExceptionType.RECLASS:
                continue
            active_entries.append(e)

        # Determine visible headers based on pipeline (P3 ID vs Cost Center)
        is_project = self.p3_id_column is not None
        
        if is_project:
            visible_headers = [
                'P3 ID', 'Accounting Period', 'WBS',
                'Document Number', 'Source Row', 'Amount', 'Type',
                'CO Doc Line Item Txt'
            ]
        else:
            visible_headers = [
                'Cost Center', 'Accounting Period', 'WBS',
                'PO/ER Number', 'Source Row', 'Amount', 'GL Line Description'
            ]

        # Style definitions matching professional sheet formatting
        hdr_fill = PatternFill("solid", fgColor="1F4E79")   # Dark blue header
        hdr_font = Font(bold=True, color="FFFFFF")          # Bold white font

        # Write header row (Row 1)
        for col_idx, h in enumerate(visible_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill

        # Write data rows
        row = 2
        for entry in active_entries:
            # Col 1: Cost Center / P3 ID
            ws.cell(row=row, column=1, value=entry.cost_center)
            # Col 2: Accounting Period
            ws.cell(row=row, column=2, value=entry.month)
            # Col 3: WBS
            ws.cell(row=row, column=3, value=entry.wbs)
            # Col 4: Document Number / PO/ER Number
            ws.cell(row=row, column=4, value=self._extract_er_number(entry.po))
            # Col 5: Source Row
            ws.cell(row=row, column=5, value=entry.row_index)
            
            # Col 6: Amount
            amt_cell = ws.cell(row=row, column=6, value=entry.amount)
            if entry.amount is not None:
                amt_cell.number_format = '#,##0.00'
                
            # Col 7 and 8 mapping based on pipeline
            if is_project:
                # Col 7: Type
                ws.cell(row=row, column=7, value=entry.transaction_type)
                # Col 8: CO Doc Line Item Txt
                ws.cell(row=row, column=8,
                        value=entry.source_row_data.get('CO Doc Line Item Txt')
                        if entry.source_row_data else None)
            else:
                # Col 7: use AP Voucher Number when no PO is attached, otherwise GL Line Description
                gl_desc = (
                    entry.source_row_data.get('GL Line Description')
                    if entry.source_row_data else None
                )
                ap_voucher = (
                    entry.source_row_data.get('AP Voucher Number')
                    if entry.source_row_data else None
                )
                ws.cell(
                    row=row,
                    column=7,
                    value=ap_voucher if not entry.po else gl_desc,
                )
            row += 1

        # Freeze Panes (Row 1 header stays frozen)
        ws.freeze_panes = "A2"

        # Auto-filter (covering ONLY the visible columns, A to H for OpEx, A to I for projects)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(visible_headers))}{row - 1}"

        # Auto-size all columns (A to H)
        for ci in range(1, len(visible_headers) + 1):
            letter = get_column_letter(ci)
            max_len = len(str(visible_headers[ci - 1]))
            # Scan down up to row-1
            for r in range(1, row):
                v = ws.cell(row=r, column=ci).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max_len + 2, 55)

    def write_exception_data_sheet(self, exception_log):
        """Write raw exception data to hidden sheet for formula reference"""
        ws = self.wb.create_sheet("Exception_Data")

        # Exclude reclasses
        class _FilteredLog:
            def __init__(self, entries):
                self.entries = entries
        exception_log = _FilteredLog([
            e for e in exception_log.entries
            if e.exception_type != ExceptionType.RECLASS
        ])

        # Headers
        headers = ['Cost Center', 'WBS', 'PO', 'Exception Type', 'Accounting Period', 'Amount', 'Type']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        
        # Data rows
        for row_idx, entry in enumerate(exception_log.entries, start=2):
            ws.cell(row=row_idx, column=1, value=entry.cost_center or '')
            ws.cell(row=row_idx, column=2, value=entry.wbs or '')
            ws.cell(row=row_idx, column=3, value=self._extract_er_number(entry.po) or '')
            ws.cell(row=row_idx, column=4, value=entry.exception_type.value)
            ws.cell(row=row_idx, column=5, value=entry.month or '')
            ws.cell(row=row_idx, column=6, value=entry.amount)
            ws.cell(row=row_idx, column=7, value=entry.transaction_type or '')
        
        # Hide the sheet
        ws.sheet_state = 'hidden'
    
    def save(self):
        """Saves the workbook to the output path."""
        try:
            self.wb.save(self.output_path)
            print(f"Workbook saved to: {self.output_path}")
        except Exception as e:
            raise IOError(
                f"Failed to save the completed workbook to the output path '{self.output_path}'. "
                f"Please ensure that the file is not currently open in Microsoft Excel, and that you have "
                f"write permissions for this directory. Internal system error: {e}"
            )
        
