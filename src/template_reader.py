from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

class TemplateReader:

    def __init__(self,
                file_path,
                header_row,
                po_col,
                po_stop_marker,
                cost_center_col,
                cost_center_start_row,
                cost_center_end_row=None,
                **kwargs
        ):
        
        
        self.wb = load_workbook(file_path)
        self.sheet: Worksheet = self.wb.active  # type: ignore[assignment]
        
        # Ensure sheet was loaded successfully
        if self.sheet is None:
            raise ValueError(f"Could not load active sheet from {file_path}")

        # Initializing instance variable
        self.header_row = header_row
        self.po_col = po_col
        self.po_stop_marker = po_stop_marker
        self.cost_center_col = cost_center_col
        self.cost_center_start_row = cost_center_start_row
        self.cost_center_end_row = cost_center_end_row

        # Read on init
        self.cost_centers = self.get_existing_cost_centers()
        self.template_rows = self.get_template_rows()
        self.pos = self.get_existing_pos()

    def _find_cost_center_start_row(self) -> int:
        """Scan column A for a 'Cost Center' title cell and return the row after it.
        Accepts any cell whose text equals 'cost center' (case-insensitive), so it
        matches 'Cost Center', 'Chargeout Cost Center', etc.
        Falls back to cost_center_start_row from config if the marker is not found.
        """
        max_row = self.sheet.max_row or 1000
        for search_row in range(1, max_row + 1):
            cell_val = self.sheet[f"{self.cost_center_col}{search_row}"].value
            if cell_val is not None and "cost center" in str(cell_val).strip().lower():
                return search_row + 1
        # Marker not found — fall back to configured start row
        print(
            f"WARNING: 'Cost Center' marker not found in column {self.cost_center_col}. "
            f"Falling back to configured cost_center_start_row={self.cost_center_start_row}."
        )
        return self.cost_center_start_row

    def get_existing_cost_centers(self) -> list[str]:
        """
        Reads cost centers from column A starting immediately after the
        'Chargeout Cost Center' header cell and stopping before any cell
        whose text is blank or starts with 'Expense'.

        The start row is auto-detected so the cost centers work regardless
        of where they appear in the sheet. Falls back to cost_center_start_row
        from config if the marker is not present.

        Returns:
            list[str]: e.g. ['1234', '2345', 'CC-999']
        """

        start_row = self._find_cost_center_start_row()
        cost_centers = []
        row = start_row

        while True:
            # Stop at end row if configured
            if self.cost_center_end_row is not None and row > self.cost_center_end_row:
                break
            cell = self.sheet[f"{self.cost_center_col}{row}"].value
            # Stop on blank or "Expense" marker
            if cell is None or str(cell).strip() == "":
                break
            cell_text = str(cell).strip()
            if cell_text.lower().startswith("expense"):
                break
            cost_center = cell_text.split("/")[0].strip()
            cost_centers.append(cost_center)
            row += 1

        if not cost_centers:
            print("WARNING: No cost centers found in template.")
        else:
            print(f"Found {len(cost_centers)} cost centers: {cost_centers}")
        return cost_centers

    def get_existing_pos(self) -> dict[str, int]:
        """Extract PO numbers and their row positions from the template."""
        pos = {po: data['row'] for po, data in self.template_rows.items()}
        self._log_pos_summary(pos)
        return pos

    def get_template_rows(self) -> dict[str, dict[str, str | int | None]]:
        """Extract front-tab PO metadata for matching transactions to template rows."""
        stop_row = self._find_stop_row()
        return self._extract_pos_from_rows(stop_row)
    
    def _find_stop_row(self) -> int:
        """Find the row containing the stop marker.
        If the marker is not present (blank / new template), returns max_row + 1
        so that _extract_pos_from_rows scans nothing and returns an empty dict
        rather than raising an error."""
        max_row = self.sheet.max_row or 1000
        for search_row in range(1, max_row + 1):
            if self.sheet[f"A{search_row}"].value == self.po_stop_marker:
                return search_row

        print(
            f"WARNING: '{self.po_stop_marker}' marker not found in template — "
            "treating as blank template with no existing POs."
        )
        return max_row + 1  # safe sentinel: loop in _extract_pos_from_rows finds nothing
    
    def _extract_pos_from_rows(self, stop_row: int) -> dict[str, dict[str, str | int | None]]:
        """Extract PO metadata from rows between header and stop marker."""
        pos = {}
        row = self.header_row + 1
        current_cost_center = None

        while row < stop_row:
            cc_value = self.sheet[f"{self.cost_center_col}{row}"].value
            if cc_value is not None:
                cc_text = str(cc_value).strip()
                if cc_text and not cc_text.lower().startswith("expense"):
                    current_cost_center = cc_text.split("/")[0].strip()

            cell_value = self.sheet[f"{self.po_col}{row}"].value

            if self._is_valid_po(cell_value):
                s = str(cell_value).strip()
                # Normalize float-formatted integers (e.g. 9500905777.0 → "9500905777")
                if s.replace('.', '', 1).replace('-', '', 1).isdigit():
                    try:
                        s = str(int(float(s)))
                    except (ValueError, OverflowError):
                        pass
                pos[s] = {
                    'row': row,
                    'cost_center': current_cost_center,
                }

            row += 1

        return pos
    
    def _is_valid_po(self, cell_value) -> bool:
        """Check if cell value represents a valid PO number."""
        if cell_value is None:
            return False
        
        po_str = str(cell_value).strip().lower()
        return po_str != "" and po_str != "none"
    
    def _log_pos_summary(self, pos: dict[str, int]) -> None:
        """Log summary of POs found in template."""
        if not pos:
            print("WARNING: No POs found in template.")
        else:
            print(f"Found {len(pos)} POs in template.")