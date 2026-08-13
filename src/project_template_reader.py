from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import re

# Month abbreviations used when scanning for metric+month header cells.
_MONTH_KWS = {
    "jan", "feb", "mar", "apr", "may", "jun",
    "jul", "aug", "sep", "oct", "nov", "dec",
    "january", "february", "march", "april", "june",
    "july", "august", "september", "october", "november", "december",
}
_METRIC_KWS = ("forecast", "accrual reversal", "accrual", "actual")


def extract_project_root(wbs: str) -> str:
    """Return the root project code from a full WBS string.

    Examples
    --------
    'CE-BTS21076'              → 'CE-BTS21076'
    'CE-BTS21076-02-10'        → 'CE-BTS21076'
    'CE-BTS21076-02-EX'        → 'CE-BTS21076'
    'CE-BTS21076-02-EX-IE'     → 'CE-BTS21076'

    Strategy: a project root is the first two dash-separated parts.
    PREFIX-PROJECTCODE[-SUB[-SUB...]] → PREFIX-PROJECTCODE
    """
    parts = wbs.strip().split('-')
    if len(parts) >= 2:
        return f"{parts[0]}-{parts[1]}"
    return wbs.strip()


def is_expense_wbs(wbs: str) -> bool:
    """Return True when the WBS code represents an expense charge.

    A WBS segment of 'EX' (case-insensitive) anywhere after the project root
    indicates an expense sub-code.

    Examples
    --------
    'CE-BTS21076-02-EX'     → True   (direct expense)
    'CE-BTS21076-02-EX-IE'  → True   (international expense)
    'CE-BTS21076-02-10'     → False  (capital)
    'CE-BTS21076'           → False  (root — not classified)
    """
    parts = wbs.strip().split('-')
    # Segments after the root (parts[0]-parts[1]) are the sub-codes
    return any(p.upper() == 'EX' for p in parts[2:])


def wbs_charge_type(wbs: str) -> str:
    """Return 'Expense' or 'Capital' for a full WBS code.

    The root code itself (only two dash-separated parts) is returned as
    'Capital' by default since it has no sub-type indicator.
    """
    return 'Expense' if is_expense_wbs(wbs) else 'Capital'


class ProjectTemplateReader:
    """Reads a project (CapEx) template.

    The template lists WBS root codes in a designated column (defaulting to
    column A) starting from a header marker row.  PO numbers are read from
    a separate column (defaulting to column B) between the header row and a
    stop marker.

    This is structurally identical to TemplateReader but uses *WBS root codes*
    as the top-level grouping key instead of cost center IDs.
    """

    def __init__(
        self,
        file_path: str,
        header_row: int,
        po_col: str,
        po_stop_marker: str,
        wbs_col: str = "A",
        p3_id_col: str | None = None,
        wbs_start_row: int = 9,
        wbs_end_row: int | None = None,
        **kwargs,
    ):
        self.wb = load_workbook(file_path)
        self.sheet: Worksheet = self.wb.active  # type: ignore[assignment]
        if self.sheet is None:
            raise ValueError(f"Could not load active sheet from {file_path}")

        # Dynamically detect the real header row; the config value is the fallback.
        self.header_row = header_row
        self.header_row = self._find_actual_header_row()
        self.po_col = po_col
        self.po_stop_marker = po_stop_marker
        self.wbs_col = wbs_col
        self.p3_id_col = p3_id_col  # Column for P3 ID (e.g., "B")
        self.wbs_start_row = wbs_start_row
        self.wbs_end_row = wbs_end_row

        # Read on init
        self.p3_wbs_map = self._get_p3_wbs_mapping()  # {p3_id: [wbs_codes]}
        # projects is derived from p3_wbs_map — the flat list of all WBS roots
        self.projects = list({
            extract_project_root(wbs)
            for wbs_list in self.p3_wbs_map.values()
            for wbs in wbs_list
        })
        self.pos = self._get_existing_pos()

    # ------------------------------------------------------------------
    # Dynamic header detection
    # ------------------------------------------------------------------

    def _find_actual_header_row(self) -> int:
        """Scan the sheet to find the real header row dynamically.

        Two strategies (same as TemplateWriter._find_actual_header_row):
        1. Look for a cell containing 'contact for po' in any column.
        2. Look for a row with ≥2 metric+month header cells
           (e.g. 'Forecast Jan', 'Actual Feb').

        Falls back to self.header_row (the config value) when neither
        signal is found.
        """
        max_col = self.sheet.max_column or 50
        max_row = self.sheet.max_row or 1000

        for r in range(1, max_row + 1):
            metric_month_hits = 0
            for c in range(1, max_col + 1):
                v = self.sheet.cell(row=r, column=c).value
                if v is None:
                    continue
                text = str(v).strip().lower()
                # Strategy 1 — explicit label
                if "contact for po" in text:
                    return r
                # Strategy 2 — metric+month header cell
                for kw in _METRIC_KWS:
                    if text.startswith(kw):
                        words = text.split()
                        if words and words[-1].rstrip('.,') in _MONTH_KWS:
                            metric_month_hits += 1
                        break
            if metric_month_hits >= 2:
                return r

        return self.header_row  # fallback to config value

    # ------------------------------------------------------------------
    # P3 ID and WBS mapping
    # ------------------------------------------------------------------

    def _get_p3_wbs_mapping(self) -> dict[str, list[str]]:
        """Read P3 IDs and their associated WBS codes from the template.
        
        Returns a mapping: {p3_id: [wbs_code1, wbs_code2, ...]}
        """
        if not self.p3_id_col:
            return {}
        
        mapping: dict[str, list[str]] = {}
        start_row = self.wbs_start_row
        
        # Find the actual start row by looking for WBS/Project header
        max_row = self.sheet.max_row or 1000
        for r in range(1, max_row + 1):
            val = self.sheet[f"{self.wbs_col}{r}"].value
            if val is None:
                continue
            low = str(val).strip().lower()
            if "wbs" in low or "project" in low:
                start_row = r + 1
                break
        
        row = start_row
        while True:
            if self.wbs_end_row is not None and row > self.wbs_end_row:
                break

            wbs_cell = self.sheet[f"{self.wbs_col}{row}"].value
            p3_cell  = self.sheet[f"{self.p3_id_col}{row}"].value

            # Stop on a completely blank WBS column only if P3 is also blank
            wbs_text = str(wbs_cell).strip() if wbs_cell is not None else ""
            p3_text  = str(p3_cell).strip()  if p3_cell  is not None else ""

            if not wbs_text and not p3_text:
                break
            if wbs_text.lower().startswith("expense"):
                break

            wbs   = wbs_text.split("/")[0].strip()
            p3_id = p3_text

            if p3_id:
                # Register the P3 ID even when there is no WBS on this row —
                # transactional rows will be matched via cost_center field directly.
                if p3_id not in mapping:
                    mapping[p3_id] = []
                if wbs and wbs not in mapping[p3_id]:
                    mapping[p3_id].append(wbs)

            row += 1

        return mapping

    # ------------------------------------------------------------------
    # Project root detection
    # ------------------------------------------------------------------

    def _find_wbs_start_row(self) -> int:
        """Scan the WBS column for a header cell containing 'wbs' or 'project'
        and return the row immediately after it.  Falls back to wbs_start_row."""
        max_row = self.sheet.max_row or 1000
        for r in range(1, max_row + 1):
            val = self.sheet[f"{self.wbs_col}{r}"].value
            if val is None:
                continue
            low = str(val).strip().lower()
            if "wbs" in low or "project" in low:
                return r + 1
        print(
            f"WARNING: WBS/Project header not found in column {self.wbs_col}. "
            f"Falling back to configured wbs_start_row={self.wbs_start_row}."
        )
        return self.wbs_start_row

    def _get_project_roots(self) -> list[str]:
        """Read WBS codes from the template and return the deduplicated list of
        project root codes (e.g. ['CE-BTS21076', 'CE-BTS22001']).

        Stops on a blank cell or a cell whose text starts with 'expense'."""
        start_row = self._find_wbs_start_row()
        seen: dict[str, None] = {}   # ordered dedup
        row = start_row

        while True:
            if self.wbs_end_row is not None and row > self.wbs_end_row:
                break
            cell = self.sheet[f"{self.wbs_col}{row}"].value
            if cell is None or str(cell).strip() == "":
                break
            text = str(cell).strip()
            if text.lower().startswith("expense"):
                break
            # Strip trailing /suffix (same convention as OpEx cost-center column)
            text = text.split("/")[0].strip()
            root = extract_project_root(text)
            if root not in seen:
                seen[root] = None
            row += 1

        projects = list(seen.keys())
        if not projects:
            print("WARNING: No project WBS codes found in template.")
        else:
            print(f"Found {len(projects)} project(s): {projects}")
        return projects

    # ------------------------------------------------------------------
    # PO / row position reading  (identical logic to TemplateReader)
    # ------------------------------------------------------------------

    def _find_stop_row(self) -> int:
        max_row = self.sheet.max_row or 1000
        for r in range(1, max_row + 1):
            if self.sheet[f"A{r}"].value == self.po_stop_marker:
                return r
        print(
            f"WARNING: '{self.po_stop_marker}' marker not found — "
            "treating as blank template with no existing POs."
        )
        return max_row + 1

    def _get_existing_pos(self) -> dict[str, int]:
        stop_row = self._find_stop_row()
        pos: dict[str, int] = {}
        row = self.header_row + 1
        while row < stop_row:
            val = self.sheet[f"{self.po_col}{row}"].value
            if val is not None:
                s = str(val).strip()
                if s and s.lower() != "none":
                    # Normalise float-formatted integers
                    if s.replace('.', '', 1).replace('-', '', 1).isdigit():
                        try:
                            s = str(int(float(s)))
                        except (ValueError, OverflowError):
                            pass
                    pos[s] = row
            row += 1
        if not pos:
            print("WARNING: No POs found in project template.")
        else:
            print(f"Found {len(pos)} POs in project template.")
        return pos
