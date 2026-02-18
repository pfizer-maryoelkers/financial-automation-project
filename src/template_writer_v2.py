from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, column_index_from_string



class TemplateWriter:

    def __init__(self, file_path: str):
        self.wb = load_workbook(file_path)
        self.sheet = self.wb.active

        # Configs for template
        self.header_row = 14    # Header row
        self.po_column = 'B'    # Col where POs are entered


        # Cost centers, WBS codes, PO numbers, and their associated rows
        self.cost_centers = {}
        self.wbs_codes = {}
        self.pos = {}

        # Column map (dynamically created starting with Dec Accrual Reversal)
        self.dec_acc_reversal_col = 'N' # Col where first data entry exists
        self.column_map = self.get_column_map()

        # Source sheet params
        self.forecast_source_cols = [
            "Forecast Fee $",
            "Forecast Expenses $",
            "Forecast Total $",
            "Total Forecast Hours 2025",
            "Jan 2025 - FTotal",
            "Feb 2025 - FTotal",
            "March 2025 - FTotal",
            "April 2025 - FTotal",
            "May 2025 - FTotal",
            "June 2025 - FTotal",
            "July 2025 - FTotal",
            "Aug 2025 - FTotal",
            "Sep 2025 - FTotal",
            "Oct 2025 - FTotal",
            "Nov 2025 - FTotal",
            "Dec 2025 - FTotal",
            "2025 Forecast total Fee"
        ]

        self.transactional_source_cols = [
            "PO Number",
            "Accounting Period",
            "AP Voucher Number",
            "Vendor Name",
            "WBS Element",
            "GL Invoice Date",
            "GL Posting Date",
            "GL Line Description",
            "Description",
            "Month",
            "GL Transaction Amount",
            "Type"
        ]



    ## Methods to get existing cost centers, WBS codes, and POs from input template sheet
    def get_existing_cost_centers(self):
        # Gets existing cost centers (if exists)
        # TODO: implement
        pass

    def get_existing_wbs(self):
        # Gets existing WBS codes (if exists)
        # TODO: implement
        pass
    
    def get_existing_pos(self):
        # Gets existing POs in the template
        row = self.header_row + 1 # Starting at row below header row

        while True:
            cell = self.sheet[f"{self.po_column}{row}"].value

            if cell is None or str(cell).strip() == "":
                break  # stop at the first blank cell

            # Normalize PO value
            po = str(cell).strip()

            # Store mapping PO → row
            self.pos[po] = row

            row += 1
    
    def get_column_map(self):
        months = [
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
        ]

        metrics = ["Accrual Reversal", "Forecast", "Accrual", "Actual"]

        col_map = {}
        col_index = column_index_from_string(self.dec_acc_reversal_col)

        for month in months:
            month_map = {}
            for metric in metrics:
                month_map[metric] = get_column_letter(col_index)
                col_index += 1

            col_map[month] = month_map

            # Skip one column before next month's block
            col_index += 1

        return col_map

    
    ## Methods to write data to sheet
    def insert_new_rows(self, num_rows):
        """
        Insert multiple rows below row 20.
        Each new row copies formulas from row 20 and clears non-formula cells.
        """
        source_row = self.header_row + 2  # Template row
        insert_at = source_row + 1  # Insert below template

        self.sheet.insert_rows(insert_at, amount=num_rows)

        # For each inserted row, copy formulas and formatting from source row
        for offset in range(num_rows):
            new_row = insert_at + offset
            for src_cell, dest_cell in zip(self.sheet[source_row], self.sheet[new_row]):
                if src_cell.has_style:
                    dest_cell._style = src_cell._style
                dest_cell.number_format = src_cell.number_format
                if src_cell.data_type == 'f':  # formula
                    translator = Translator(src_cell.value, origin=src_cell.coordinate)
                    dest_cell.value = translator.translate_formula(dest_cell.coordinate)
                else:
                    dest_cell.value = None  # Keep new row blank except formulas

        print(f"Inserted {num_rows} blank rows.\n")
        

    def write_data(self, data:dict):
        '''
        Writes data to template.

        Expects data dict in format:

            dict: {
                'PO12345': {
                    'Jan': {
                        'Forecast': 1050,
                        'Actual': 900
                        'Accrual': 950.0,
                        'Accrual Reversal': 0.0,
                    },
                    'Feb': {
                        'Forecast': 950,
                        'Actual': 800
                        'Accrual': 1050.0,
                        'Accrual Reversal': -950,
                    }
                    ...
                }
                ...
            }

        '''
        num_pos = len(self.pos)
        extra_rows = max(0, num_pos - 2)

        # Pre-allocate extra rows if needed
        if extra_rows > 0:
            self.insert_new_rows(extra_rows)


        for po, row in self.pos.items():

            # Always write the PO itself
            self.sheet[f"{self.po_column}{row}"] = po

            # If PO is not found in the data → leave row blank and print
            if po not in data:
                print(f"PO '{po}' found in template but not in source data. Leaving blank.")
                continue

            # Write PO into the sheet (if the blank template expects it)
            self.sheet[f"{self.po_column}{row}"] = po

            # Loop through months for this PO
            for month, metrics in data[po].items():
                if month not in self.column_map:
                    continue  # ignore unexpected months

                # Loop through metrics (Forecast, Actual, Accrual, Accrual Reversal)
                for metric, col_letter in self.column_map[month].items():
                    value = metrics.get(metric)

                    # Write the value (could be None or numeric)
                    self.sheet[f"{col_letter}{row}"] = value



    ## Methods to write source sheets
    def write_forecast_source_sheet(self, forecast_df):
        # Method to write forecast source sheet. 
        # Filter to template POs
        pos_list = list(self.pos.keys())
        df = forecast_df[forecast_df["PO #"].astype(str).isin(pos_list)].copy()

        # Visible first, hidden second
        visible = [c for c in self.forecast_source_cols if c in df.columns]
        hidden = [c for c in df.columns if c not in visible]
        final_cols = visible + hidden
        df = df[final_cols]

        ws = self.wb.create_sheet("Forecast Source Data")

        # Header
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(1, col_idx, col_name)

        # Data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row_idx, col_idx, value)

        # Totals row
        start = 2
        end = ws.max_row
        total_row = end + 1
        ws.cell(total_row, 1, "PO Total")

        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name not in hidden:
                continue
            col_letter = get_column_letter(col_idx)
            ws.cell(
                total_row,
                col_idx,
                f"=SUBTOTAL(9, {col_letter}{start}:{col_letter}{end})"
            )

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        # Auto width
        for col_idx, col_name in enumerate(df.columns, 1):
            letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for r in range(2, end + 1):
                v = ws.cell(r, col_idx).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = max_len + 2

        # Group hidden columns
        if hidden:
            start_col = len(visible) + 1
            end_col = len(final_cols)
            ws.column_dimensions.group(
                get_column_letter(start_col),
                get_column_letter(end_col),
                hidden=True
            )

    def write_transactional_source_sheet(self, transactions_df):
        # Method to write transactional detail source sheet
        # Filter to POs present in the template
        pos_list = list(self.pos.keys())
        df = transactions_df[transactions_df["PO Number"].astype(str).isin(pos_list)].copy()

        # Build final column ordering
        visible = [c for c in self.transactional_source_cols if c in df.columns]
        hidden = [c for c in df.columns if c not in visible]
        final_cols = visible + hidden
        df = df[final_cols]

        ws = self.wb.create_sheet("Transactions Source Data")

        # Header
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(1, col_idx, col_name)

        # Data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row_idx, col_idx, value)

        if visible:
            last_vis_col = len(visible)
            last_letter = get_column_letter(last_vis_col)
            ws.auto_filter.ref = f"A1:{last_letter}1"
        else:
            ws.auto_filter.ref = ws.dimensions

        ws.freeze_panes = "A2"

        # Auto-size columns
        end = ws.max_row
        for col_idx, col_name in enumerate(df.columns, 1):
            letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for row_idx in range(2, end + 1):
                val = ws.cell(row_idx, col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = max_len + 2

        # Collapse hidden columns (group them)
        if hidden:
            start = len(visible) + 1
            end = len(final_cols)
            ws.column_dimensions.group(
                get_column_letter(start),
                get_column_letter(end),
                hidden=True
            )
